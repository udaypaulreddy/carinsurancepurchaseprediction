from django.db.models import Count
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
import datetime
import openpyxl


# Create your views here.
from Remote_User.models import ClientRegister_Model,Car_Insurance,Car_Insurance_Prediction,detection_ratio,detection_accuracy

def login(request):


    if request.method == "POST" and 'submit1' in request.POST:

        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            enter = ClientRegister_Model.objects.get(username=username,password=password)
            request.session["userid"] = enter.id

            return redirect('Add_DataSet_Details')
        except:
            pass

    return render(request,'RUser/login.html')

def Add_DataSet_Details(request):
    if "GET" == request.method:
        return render(request, 'RUser/Add_DataSet_Details.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        #print(sheets)
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        #print(worksheet)
        # getting active sheet
        active_sheet = wb.active
        #print(active_sheet)
        # reading a cell
        #print(worksheet["A1"].value)
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                #print(cell.value)
            excel_data.append(row_data)
            Car_Insurance.objects.all().delete()
    for r in range(1, active_sheet.max_row + 1):
        Car_Insurance.objects.create(
        idno=active_sheet.cell(r, 1).value,
        Gender=active_sheet.cell(r, 2).value,
        Age=active_sheet.cell(r, 3).value,
        Driving_License=active_sheet.cell(r, 4).value,
        Region_Code=active_sheet.cell(r, 5).value,
        Previously_Insured=active_sheet.cell(r, 6).value,
        Vehicle_Age=active_sheet.cell(r, 7).value,
        Vehicle_Damage=active_sheet.cell(r, 8).value,
        Annual_Premium=active_sheet.cell(r, 9).value,
        Policy_Sales_Channel=active_sheet.cell(r, 10).value,
        Vintage=active_sheet.cell(r, 11).value,
        IResponse=active_sheet.cell(r, 12).value
        )

    return render(request, 'RUser/Add_DataSet_Details.html', {"excel_data": excel_data})


def Register1(request):

    if request.method == "POST":
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phoneno = request.POST.get('phoneno')
        country = request.POST.get('country')
        state = request.POST.get('state')
        city = request.POST.get('city')
        ClientRegister_Model.objects.create(username=username, email=email, password=password, phoneno=phoneno,
                                            country=country, state=state, city=city)

        return render(request, 'RUser/Register1.html')
    else:
        return render(request,'RUser/Register1.html')

def ViewYourProfile(request):
    userid = request.session['userid']
    obj = ClientRegister_Model.objects.get(id= userid)
    return render(request,'RUser/ViewYourProfile.html',{'object':obj})


def predict_car_Insurance_purchase(request):
    predict1=""
    if request.method == "POST":

        if request.method == "POST":
            Gender = request.POST.get('Gender')
            Age = request.POST.get('Age')
            Dl = request.POST.get('Dl')
            Rc = request.POST.get('Rc')
            Pi = request.POST.get('Pi')
            Vage = request.POST.get('Vage')
            Vd = request.POST.get('Vd')
            Ap = request.POST.get('Ap')
            Psc = request.POST.get('Psc')
            Vi = request.POST.get('Vi')

            obj = Car_Insurance.objects.get(Gender=Gender,Age=Age,Driving_License=Dl,Region_Code=Rc,Previously_Insured=Pi,Vehicle_Age=Vage,Vehicle_Damage=Vd,Annual_Premium=Ap,Policy_Sales_Channel=Psc,Vintage=Vi )

            val=int(obj.IResponse)

            print("Value")
            print(val)

            if val==1:
                predict1="Insurance Purchase Interested"
            elif val==0:
                predict1 = "Insurance Purchase Not Interested"

            print(predict1)

            Car_Insurance_Prediction.objects.create(Gender=Gender, Age=Age, Driving_License=Dl, Region_Code=Rc, Previously_Insured=Pi,
                                      Vehicle_Age=Vage, Vehicle_Damage=Vd, Annual_Premium=Ap, Policy_Sales_Channel=Psc,
                                      Vintage=Vi,IPrediction=predict1)


        return render(request, 'RUser/predict_car_Insurance_purchase.html',{'objs': predict1})
    return render(request, 'RUser/predict_car_Insurance_purchase.html')



